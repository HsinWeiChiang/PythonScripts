import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime, timedelta




def getTX(Date_ID):
    #期交所 
    url = 'https://www.taifex.com.tw/cht/3/futAndOptDate'

    # Post 資訊，
    form_data = {
        'queryType': '1',
        'goDay': '',
        'dateaddcnt': '-1',
        'doQuery': '1',
        'queryDate': Date_ID
    }


   # 发送HTTP请求获取网站的HTML内容
    response = requests.post(url, params= form_data )

    # 检查响应状态码
    if response.status_code == 200:
        # 使用BeautifulSoup解析HTML
        soup = BeautifulSoup(response.text, 'html.parser')        
        tr_element = soup.find_all('tr', class_='12bk')
        
        # 解析表格資料並存儲在列表中
        data = []
        for tr in tr_element:
            tds = tr.find_all('td')
            tds_data = []
            
            for td in tds:
                tds_data.append(td.text.strip())
            data.append(tds_data)     

        real_data = [sublist for sublist in data if sublist != [''] and sublist]
        
        keys = ["自營商", "投信", "外資", "合計"]
        data_with_keys = [[keys[i]] + sublist for i, sublist in enumerate(real_data)]
        
        ID = ["身份別","期貨","選擇權","期貨","選擇權","期貨","選擇權","期貨","選擇權","期貨","選擇權","期貨","選擇權"]
        data_with_keys.insert(0, ID)

        lotsMoney = ["","口數","口數","契約金額","契約金額","口數","口數","契約金額","契約金額","口數","口數","契約金額","契約金額"]
        data_with_keys.insert(0, lotsMoney)
        
        Bull_Bear = ["","多方","多方","多方","多方","空方","空方","空方","空方","多空淨額","多空淨額","多空淨額","多空淨額"]
        data_with_keys.insert(0, Bull_Bear)

        time_tag = [date_str]  # 顯示日期所對應抓取資料
        data_with_keys.insert(0, time_tag)

        print(data_with_keys)
        
        
        
        # 檢查是否excel分頁
        file_path = 'test2.xlsx'
        sheet_name = '三大法人當日交易口數與金額'
        new_data = data_with_keys
        
        append_data_to_sheet(file_path, sheet_name, new_data)
        

    else:
        print("请求失败，状态码：", response.status_code)


def append_data_to_sheet(file_path, sheet_name, new_data):
        # 加载现有的工作簿
        wb = load_workbook(file_path)
        
        # 获取指定的工作表，如果不存在则创建
        sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        
        # 在工作表中找到第一个空行
        row_number = sheet.max_row + 1

        # 将新数据写入工作表
        for row in new_data:
            sheet.append(row)

        # 保存工作簿
        wb.save(file_path)
        


if __name__ == "__main__":

    # 定義開始日期和結束日期
    start_date = datetime.strptime('2024/03/29', '%Y/%m/%d')
    end_date = datetime.strptime('2024/04/03', '%Y/%m/%d')
    
    # 初始化日期字串列表
    date_strings = []
    
    # 遍歷日期範圍，並將每個日期轉換為字串並添加到列表中
    current_date = start_date
    
    while current_date <= end_date:
        date_strings.append(current_date.strftime('%Y/%m/%d'))
        current_date += timedelta(days=1)
    
    for date_str in date_strings:
        
        print(date_str)
        #執行當下時間的網頁抓取 
        getTX(date_str)
        


